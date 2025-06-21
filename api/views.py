from rest_framework import viewsets, permissions
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from django.db import IntegrityError
from rest_framework.exceptions import ValidationError
from .models import (
    KhachHang, Sach, HoaDon, CT_HoaDon, PhieuThuTien,
    DauSach, TacGia, TheLoai, NXB,
    PhieuNhapSach, CT_NhapSach,
    BaoCaoTon, CT_BCTon, BaoCaoCongNo, CT_BCCongNo, ThamSo
)

from .serializers import (
    KhachHangSerializer, SachSerializer, HoaDonSerializer, CTHoaDonSerializer, PhieuThuTienSerializer,
    DauSachSerializer, TacGiaSerializer, TheLoaiSerializer, NXBSerializer,
    PhieuNhapSachSerializer, CTNhapSachSerializer,
    BaoCaoTonSerializer, CT_BCTonSerializer, BaoCaoCongNoSerializer, CT_BCCNSerializer, ThamSoSerializer,
    GroupSerializer    
)

from django_filters.rest_framework import DjangoFilterBackend, FilterSet, NumberFilter, CharFilter, DateFilter
from django.contrib.auth.models import User, Group
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response
from .serializers import UserSerializer, CreateUserSerializer
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import Paragraph
from django.conf import settings
import os
from reportlab.lib.utils import ImageReader


import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from django.http import HttpResponse
from openpyxl.utils import get_column_letter


def get_valid_groups():
    """Get all group names from the database"""
    return list(Group.objects.values_list('name', flat=True))
# class IsAuthenticated(permissions.BasePermission):
#     def has_permission(self, request, view):
#         if not request.user or not request.user.is_authenticated:
#             return False

#         if request.user.is_superuser:
#             return True

#         model_name = view.queryset.model._meta.model_name
#         user_groups = request.user.groups.all()
        
#         group_perms = GroupModelPermission.objects.filter(
#             group__in=user_groups,
#             model_name=model_name
#         )

#         if request.method in permissions.SAFE_METHODS:
#             return group_perms.filter(can_view=True).exists()
#         elif request.method == 'POST':
#             return group_perms.filter(can_add=True).exists()
#         elif request.method in ['PUT', 'PATCH']:
#             return group_perms.filter(can_change=True).exists()
#         elif request.method == 'DELETE':
#             return group_perms.filter(can_delete=True).exists()

class UserManagementViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]  # Any authenticated user can access

    @action(detail=False, methods=['post'])
    def create_staff(self, request):
        serializer = CreateUserSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()  # This will handle role assignment
            # Set gender if provided
            if 'gioiTinh' in request.data:
                user.profile.gioiTinh = request.data['gioiTinh']
                user.profile.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def change_password(self, request, pk=None):
        user = self.get_object()
        new_password = request.data.get('new_password')
        if new_password:
            user.set_password(new_password)
            user.save()
            return Response({"status": "password changed"})
        return Response(
            {"error": "No password provided"}, 
            status=status.HTTP_400_BAD_REQUEST
        )

    @action(detail=True, methods=['put'])
    def update_profile(self, request, pk=None):
        user = self.get_object()
        serializer = UserSerializer(user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def deactivate(self, request, pk=None):
        user = self.get_object()
        user.is_active = False
        user.save()
        return Response({"status": "user deactivated"})
    @action(detail=True, methods=['post'])
    def reactivate(self, request, pk=None):
        user = self.get_object()
        user.is_active = True
        user.save()
        return Response({"status": "user reactivated"})

# CT_HoaDon: Nguoilaphd (full), Quanli (all)
class CTHoaDonViewSet(viewsets.ModelViewSet):
    queryset = CT_HoaDon.objects.all()
    serializer_class = CTHoaDonSerializer

    permission_classes = [IsAuthenticated]

# PhieuThuTien: Nguoithu (full), Quanli (all)
class PhieuThuTienViewSet(viewsets.ModelViewSet):
    queryset = PhieuThuTien.objects.all()
    serializer_class = PhieuThuTienSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request, pk=None):
        try:
            # Register fonts
            font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'arial unicode ms.otf')
            font_bold_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'arial unicode ms bold.otf')
            pdfmetrics.registerFont(TTFont('Arial Unicode', font_path))
            pdfmetrics.registerFont(TTFont('Arial Unicode Bold', font_bold_path))

            # Fetch data
            phieuthu = self.get_object()

            # Create PDF response
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="phieuthu_{phieuthu.MaPhieuThu}.pdf"'

            # Set up PDF with margins
            pdf = canvas.Canvas(response, pagesize=A4)
            width, height = A4
            left_margin = 2*cm
            right_margin = width - 2*cm
            effective_width = right_margin - left_margin

            # Add logo
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'logo.png')
            pdf.drawImage(logo_path, left_margin, height - 3*cm, width=50, height=50)

            # Title with more spacing
            pdf.setFont('Arial Unicode Bold', 16)
            pdf.drawString(left_margin + (effective_width/2 - 70), height - 3.5*cm, "CHI TIẾT PHIẾU THU TIỀN")

            # Start content below logo and title
            y = height - 5*cm

            # Two columns for information
            col_width = (effective_width - 1*cm) / 2
            
            # Information headers with blue background
            pdf.setFillColor(colors.HexColor('#4D94FF'))
            pdf.rect(left_margin, y, col_width, 25, fill=1)
            pdf.rect(left_margin + col_width + 1*cm, y, col_width, 25, fill=1)
            
            # Header text
            pdf.setFillColor(colors.white)
            pdf.setFont('Arial Unicode Bold', 12)
            pdf.drawString(left_margin + 50, y + 7, "Thông Tin Phiếu Thu")
            pdf.drawString(left_margin + col_width + 1*cm + 50, y + 7, "Thông Tin Khách Hàng")

            # Content
            pdf.setFillColor(colors.black)
            pdf.setFont('Arial Unicode', 12)
            y -= 40

            # Left column content
            left_x = left_margin
            pdf.drawString(left_x, y, f"Ngày Thu: {phieuthu.NgayThu.strftime('%d/%m/%Y')}")
            pdf.drawString(left_x, y - 25, f"Mã Phiếu Thu: PT{phieuthu.MaPhieuThu:03d}")
            pdf.drawString(left_x, y - 50, f"Mã Nhân Viên: NV{phieuthu.NguoiThu.id:03d}")
            pdf.drawString(left_x, y - 75, f"Họ Tên: {phieuthu.NguoiThu.first_name} {phieuthu.NguoiThu.last_name}")

            # Right column content
            right_x = left_margin + col_width + 1*cm
            pdf.drawString(right_x, y, f"Mã Khách: KH{phieuthu.MaKH.MaKhachHang:03d}")
            pdf.drawString(right_x, y - 25, f"Họ Tên: {phieuthu.MaKH.HoTen}")
            pdf.drawString(right_x, y - 50, f"Số Điện Thoại: {phieuthu.MaKH.DienThoai}")
            pdf.drawString(right_x, y - 75, f"Email: {phieuthu.MaKH.Email}")

            # Amount information
            y -= 125
            pdf.drawString(left_margin, y, f"Số Tiền Thu: {phieuthu.SoTienThu:,} VNĐ")

            pdf.save()
            return response

        except Exception as e:
            return Response({"detail": str(e)}, status=400)

# DauSach: Nguoi nhap (full), Quanli (all)
class DauSachViewSet(viewsets.ModelViewSet):
    queryset = DauSach.objects.all()
    serializer_class = DauSachSerializer
    permission_classes = [IsAuthenticated]

# TacGia: Nguoi nhap (full), Quanli (all)
class TacGiaViewSet(viewsets.ModelViewSet):
    queryset = TacGia.objects.all()
    serializer_class = TacGiaSerializer
    permission_classes = [IsAuthenticated]

class NXBViewSet(viewsets.ModelViewSet):
    queryset = NXB.objects.all()
    serializer_class = NXBSerializer
    permission_classes = [IsAuthenticated]

# TheLoai: Nguoi nhap (full), Quanli (all)
class TheLoaiViewSet(viewsets.ModelViewSet):
    queryset = TheLoai.objects.all()
    serializer_class = TheLoaiSerializer
    permission_classes = [IsAuthenticated]

# CT_NhapSach: Nguoi nhap (full), Quanli (all)
class CTNhapSachViewSet(viewsets.ModelViewSet):
    queryset = CT_NhapSach.objects.all()
    serializer_class = CTNhapSachSerializer

    permission_classes = [IsAuthenticated]

class CTBCTonViewSet(viewsets.ModelViewSet):
    queryset = CT_BCTon.objects.all()
    serializer_class = CT_BCTonSerializer
    permission_classes = [IsAuthenticated]

class CTBCCongNoViewSet(viewsets.ModelViewSet):
    queryset = CT_BCCongNo.objects.all()
    serializer_class = CT_BCCNSerializer
    permission_classes = [IsAuthenticated]

class ThamSoViewSet(viewsets.ModelViewSet):
    queryset = ThamSo.objects.all()
    serializer_class = ThamSoSerializer
    permission_classes = [IsAuthenticated]

# Custom filter for Sach
class SachFilter(FilterSet):
    # Filter by book name (DauSach.TenSach)
    ten_sach = CharFilter(field_name='MaDauSach__TenSach', lookup_expr='icontains')
    # Filter by author name (TacGia.TenTG, many-to-many)
    tac_gia = CharFilter(method='filter_tac_gia')
    # Filter by genre name (TheLoai.TenTheLoai)
    the_loai = CharFilter(field_name='MaDauSach__MaTheLoai__TenTheLoai', lookup_expr='icontains')
    # Filter by publisher (Sach.NXB)
    nxb = CharFilter(field_name='NXB__TenNXB', lookup_expr='icontains')
    # Filter by year (Sach.NamXB)
    nam_xb = NumberFilter(field_name='NamXB')
    nam_xb_min = NumberFilter(field_name='NamXB', lookup_expr='gte')
    nam_xb_max = NumberFilter(field_name='NamXB', lookup_expr='lte')
    # Filter by stock (Sach.SLTon)
    slton_min = NumberFilter(field_name='SLTon', lookup_expr='gte')
    slton_max = NumberFilter(field_name='SLTon', lookup_expr='lte')

    def filter_tac_gia(self, queryset, name, value):
        # MaDauSach is FK to DauSach, MaTG is M2M in DauSach
        return queryset.filter(MaDauSach__MaTG__TenTG__icontains=value)

    class Meta:
        model = Sach
        fields = [
            'ten_sach', 'tac_gia', 'the_loai', 'nxb',
            'nam_xb', 'nam_xb_min', 'nam_xb_max',
            'slton_min', 'slton_max'
        ]

class SachViewSet(viewsets.ModelViewSet):
    queryset = Sach.objects.all()
    serializer_class = SachSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = SachFilter

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        # Kiểm tra có chi tiết hóa đơn hoặc phiếu nhap không
        has_invoice = CT_HoaDon.objects.filter(MaSach=instance).exists()
        has_nhapsach = CT_NhapSach.objects.filter(MaSach=instance).exists()
        if has_invoice or has_nhapsach:
            raise ValidationError({'detail': 'Không thể xóa sách vì đã có hóa đơn hoặc phiếu nhập liên quan.'})
        # Kiểm tra các dòng công nợ liên quan
        ct_bcton_list = CT_BCTon.objects.filter(MaSach=instance)
        if ct_bcton_list.exists():
            all_zero = all(
                (getattr(ct, 'TonDau', 0) == 0 and getattr(ct, 'PhatSinh', 0) == 0 and getattr(ct, 'TonCuoi', 0) == 0)
                for ct in ct_bcton_list
            )
            if all_zero:
                ct_bcton_list.delete()
            else:
                raise ValidationError({'detail': 'Không thể xóa sách vì còn dòng báo cáo tồn có số dư khác 0.'})
        return super().destroy(request, *args, **kwargs)

class KhachHangFilter(FilterSet):
    id = NumberFilter(field_name='MaKhachHang')
    name = CharFilter(field_name='HoTen', lookup_expr='icontains')
    phone = CharFilter(field_name='DienThoai', lookup_expr='icontains')
    invoiceId = CharFilter(field_name='hoadon__MaHD', lookup_expr='icontains')  # Ensure 'hoadon' is the related_name for HoaDon in KhachHang

    class Meta:
        model = KhachHang
        fields = ['MaKhachHang', 'HoTen', 'DienThoai', 'hoadon__MaHD']

class KhachHangViewSet(viewsets.ModelViewSet):
    queryset = KhachHang.objects.all()
    serializer_class = KhachHangSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = KhachHangFilter

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        from .models import CT_BCCongNo, HoaDon, PhieuThuTien
        # Kiểm tra có hóa đơn hoặc phiếu thu không
        has_invoice = HoaDon.objects.filter(MaKH=instance).exists()
        has_receipt = PhieuThuTien.objects.filter(MaKH=instance).exists()
        if has_invoice or has_receipt:
            raise ValidationError({'detail': 'Không thể xóa khách hàng vì đã có hóa đơn hoặc phiếu thu liên quan.'})
        # Kiểm tra các dòng công nợ liên quan
        ct_bccn_list = CT_BCCongNo.objects.filter(MaKH=instance)
        if ct_bccn_list.exists():
            all_zero = all(
                (getattr(ct, 'NoDau', 0) == 0 and getattr(ct, 'PhatSinh', 0) == 0 and getattr(ct, 'NoCuoi', 0) == 0)
                for ct in ct_bccn_list
            )
            if all_zero:
                ct_bccn_list.delete()
            else:
                raise ValidationError({'detail': 'Không thể xóa khách hàng vì còn dòng công nợ có số dư khác 0.'})
        return super().destroy(request, *args, **kwargs)

# PhieuNhapSach filterset
class PhieuNhapSachFilter(FilterSet):
    maNhap = NumberFilter(field_name='MaPhieuNhap')
    ngayNhap = DateFilter(field_name='NgayNhap')
    # maSach and tenSach are not direct fields, but you can filter by related Sach
    maSach = NumberFilter(field_name='MaSach')
    tenSach = CharFilter(field_name='MaSach__MaDauSach__TenSach', lookup_expr='icontains')

    class Meta:
        model = PhieuNhapSach
        fields = ['MaPhieuNhap', 'NgayNhap', 'MaSach', 'MaSach__MaDauSach__TenSach']

# PhieuNhapSach: Nguoi nhap (full), Quanli (all)
class PhieuNhapSachViewSet(viewsets.ModelViewSet):
    queryset = PhieuNhapSach.objects.all()
    serializer_class = PhieuNhapSachSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = PhieuNhapSachFilter

# BaoCaoTon filterset
class BaoCaoTonFilter(FilterSet):
    thang = DateFilter(field_name='Thang')
    # book/stock: filter by related CT_BCTon
    maSach = NumberFilter(field_name='MaSach__MaSach')
    tenSach = CharFilter(field_name='MaSach__MaDauSach__TenSach', lookup_expr='icontains')
    # tonDau = NumberFilter(field_name='CT_BCTon__TonDau')
    # tonCuoi = NumberFilter(field_name='CT_BCTon__TonCuoi')

    class Meta:
        model = BaoCaoTon
        fields = ['Thang', 'MaSach__MaSach', 'MaSach__MaDauSach__TenSach']

class BaoCaoTonViewSet(viewsets.ModelViewSet):
    queryset = BaoCaoTon.objects.all()
    serializer_class = BaoCaoTonSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = BaoCaoTonFilter

# BaoCaoCongNo filterset
class BaoCaoCongNoFilter(FilterSet):
    thang = DateFilter(field_name='Thang')
    maKH = NumberFilter(field_name='MaKH__MaKhachHang')
    tenKH = CharFilter(field_name='MaKH__HoTen', lookup_expr='icontains')
    # noDau = NumberFilter(field_name='CT_BCCN__NoDau')
    # noCuoi = NumberFilter(field_name='CT_BCCN__NoCuoi')

    class Meta:
        model = BaoCaoCongNo
        fields = ['Thang', 'MaKH__MaKhachHang', 'MaKH__HoTen']

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def bcton_excel(request, bcton_id):
    try:
        bct = BaoCaoTon.objects.get(MaBCTon=bcton_id)
    except BaoCaoTon.DoesNotExist:
        return HttpResponse("Báo cáo tồn không tồn tại", status=404)

    # Lấy danh sách chi tiết tồn
    ct_list = CT_BCTon.objects.filter(MaBCTon=bct).select_related('MaSach')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Báo cáo tồn - Tháng {bct.Thang.strftime('%m-%Y')}"

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Header
    ws.merge_cells('A1:H1')  # Merge cells across 6 columns
    title_cell = ws['A1']
    title_cell.value = "BÁO CÁO TỒN TIỆM SÁCH TRÂN TRÂN"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:H2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Tháng {bct.Thang.strftime('%m-%Y')}"
    subtitle_cell.font = Font(size=12, bold=True)
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.append([])

    # Table header
    headers = ["No.", "Mã sách", "Tên sách", "NXB", "Năm xuất bản", "Tồn đầu", "Phát sinh", "Tồn cuối"]
    ws.append(headers)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # Table rows
    for idx, ct in enumerate(ct_list, start=1):
        row_data = [
            idx,
            f"S{ct.MaSach.MaSach:03d}",
            ct.MaSach.MaDauSach.TenSach,
            ct.MaSach.NXB.TenNXB,
            ct.MaSach.NamXB,
            ct.TonDau,
            ct.PhatSinh,
            ct.TonCuoi
        ]
        ws.append(row_data)

        # Apply borders and alignment to the entire row
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=4 + idx, column=col_num)
            cell.alignment = center_align
            cell.border = thin_border

    # Auto-size columns
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"baocaoton_{bct.Thang.strftime('%m_%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def bccno_excel(request, bccno_id):
    try:
        bccn = BaoCaoCongNo.objects.get(MaBCCN=bccno_id)
    except BaoCaoCongNo.DoesNotExist:
        return HttpResponse("Báo cáo công nợ không tồn tại", status=404)

    # Lấy danh sách chi tiết tồn
    ct_list = CT_BCCongNo.objects.filter(MaBCCN=bccn).select_related('MaKH')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Báo cáo công nợ - Tháng {bccn.Thang.strftime('%m-%Y')}"

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Header
    ws.merge_cells('A1:H1')  # Merge cells across 6 columns
    title_cell = ws['A1']
    title_cell.value = "BÁO CÁO CÔNG NỢ TIỆM SÁCH TRÂN TRÂN"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:H2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Tháng {bccn.Thang.strftime('%m-%Y')}"
    subtitle_cell.font = Font(size=12, bold=True)
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.append([])

    # Table header
    headers = ["No.", "Mã khách hàng", "Tên khách hàng", "Số Điện Thoại", "Email", "Nợ đầu", "Phát sinh", "Nợ cuối"]
    ws.append(headers)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # Table rows
    for idx, ct in enumerate(ct_list, start=1):
        row_data = [
            idx,
            f"KH{ct.MaKH.MaKhachHang:03d}",
            ct.MaKH.HoTen,
            ct.MaKH.DienThoai,
            ct.MaKH.Email,
            ct.NoDau,
            ct.PhatSinh,
            ct.NoCuoi
        ]
        ws.append(row_data)

        # Apply borders and alignment to the entire row
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=4 + idx, column=col_num)
            cell.alignment = center_align
            cell.border = thin_border

    # Auto-size columns
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"baocaocongno_{bccn.Thang.strftime('%m_%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

class BaoCaoCongNoViewSet(viewsets.ModelViewSet):
    queryset = BaoCaoCongNo.objects.all()
    serializer_class = BaoCaoCongNoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = BaoCaoCongNoFilter    

# HoaDon filterset
class HoaDonFilter(FilterSet):
    maHD = NumberFilter(field_name='MaHD')
    ngayLap = DateFilter(field_name='NgayLap')
    maKH = NumberFilter(field_name='MaKH__MaKhachHang')
    tenKH = CharFilter(field_name='MaKH__HoTen', lookup_expr='icontains')
    # status is not a field in HoaDon, so skip it

    class Meta:
        model = HoaDon
        fields = ['MaHD', 'NgayLap', 'MaKH__MaKhachHang', 'MaKH__HoTen']
# HoaDon: Nguoilaphd (full), Nguoithu (view), Quanli (all)

from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import cm
from reportlab.lib import colors
import os

class HoaDonViewSet(viewsets.ModelViewSet):
    queryset = HoaDon.objects.all()
    serializer_class = HoaDonSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = HoaDonFilter

    @action(detail=True, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request, pk=None):
        try:
            # Register fonts
            font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'arial unicode ms.otf')
            font_bold_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'arial unicode ms bold.otf')
            pdfmetrics.registerFont(TTFont('Arial Unicode', font_path))
            pdfmetrics.registerFont(TTFont('Arial Unicode Bold', font_bold_path))

            # Fetch data
            hoadon = self.get_object()
            cthoadon_list = CT_HoaDon.objects.filter(MaHD=hoadon)

            # Create PDF response
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="hoadon_{hoadon.MaHD}.pdf"'

            # Set up PDF with margins
            pdf = canvas.Canvas(response, pagesize=A4)
            width, height = A4
            left_margin = 2*cm
            right_margin = width - 2*cm
            effective_width = right_margin - left_margin

            # Add logo
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'logo.png')
            pdf.drawImage(logo_path, left_margin, height - 3*cm, width=50, height=50)

            # Title with more spacing
            pdf.setFont('Arial Unicode Bold', 16)
            pdf.drawString(left_margin + (effective_width/2 - 70), height - 3.5*cm, "CHI TIẾT HÓA ĐƠN")

            # Start content below logo and title
            y = height - 5*cm

            # Two columns for information
            col_width = (effective_width - 1*cm) / 2
            
            # Information headers with blue background
            pdf.setFillColor(colors.HexColor('#4D94FF'))
            pdf.rect(left_margin, y, col_width, 25, fill=1)
            pdf.rect(left_margin + col_width + 1*cm, y, col_width, 25, fill=1)
            
            # Header text
            pdf.setFillColor(colors.white)
            pdf.setFont('Arial Unicode Bold', 12)
            pdf.drawString(left_margin + 50, y + 7, "Thông Tin Hóa Đơn")
            pdf.drawString(left_margin + col_width + 1*cm + 50, y + 7, "Thông Tin Khách Hàng")

            # Content
            pdf.setFillColor(colors.black)
            pdf.setFont('Arial Unicode', 12)
            y -= 40

            # Left column content
            left_x = left_margin
            pdf.drawString(left_x, y, f"Ngày Lập: {hoadon.NgayLap.strftime('%d/%m/%Y')}")
            pdf.drawString(left_x, y - 25, f"Mã Hóa Đơn: HD{hoadon.MaHD:03d}")
            pdf.drawString(left_x, y - 50, f"Mã Nhân Viên: NV{hoadon.NguoiLapHD.id:03d}")
            pdf.drawString(left_x, y - 75, f"Tên Nhân Viên: {hoadon.NguoiLapHD.first_name} {hoadon.NguoiLapHD.last_name}")

            # Right column content
            right_x = left_margin + col_width + 1*cm
            pdf.drawString(right_x, y, f"Mã Khách Hàng: KH{hoadon.MaKH.MaKhachHang:03d}")
            pdf.drawString(right_x, y - 25, f"Tên Khách Hàng: {hoadon.MaKH.HoTen}")
            pdf.drawString(right_x, y - 50, f"Số Điện Thoại: {hoadon.MaKH.DienThoai}")
            pdf.drawString(right_x, y - 75, f"Email: {hoadon.MaKH.Email}")

            # Book list header
            y -= 125
            pdf.setFillColor(colors.HexColor('#4D94FF'))
            pdf.rect(left_margin, y, effective_width, 25, fill=1)
            
            pdf.setFillColor(colors.white)
            pdf.setFont('Arial Unicode Bold', 12)
            pdf.drawString(left_margin + (effective_width/2 - 80), y + 7, "Danh Sách Sách Đã Mua")
            
            pdf.setFillColor(colors.black)
            # Table headers
            y -= 30
            headers = ['No.', 'Mã Sách', 'Tên Sách', 'Số Lượng', 'Đơn Giá', 'Thành Tiền']
            col_widths = [30, 60, effective_width - 350, 60, 80, 120]  # Last column wider
            
            # Draw header cells with background
            x = left_margin
           
            # pdf.setFillColor(colors.HexColor('#E5E5E5'))  # Light gray background for headers
            for header, width in zip(headers, col_widths):
                pdf.rect(x, y, width, 25, fill=0)
                # Center align header text
                text_width = pdf.stringWidth(header, 'Arial Unicode Bold', 12)
                # pdf.setFillColor(colors.black)
                pdf.drawString(x + (width - text_width)/2, y + 7, header)
                x += width

            # Table content with word wrapping for book title
            def draw_wrapped_text(text, x, y, width, height):
                from reportlab.lib.utils import simpleSplit
                text_lines = simpleSplit(text, 'Arial Unicode', 12, width - 10)
                y_offset = height - (row_height - 20)/2
                for line in text_lines:
                    pdf.drawString(x + 5, y + y_offset, line)
                    y_offset -= 15

            # Content rows
            pdf.setFont('Arial Unicode', 12)
            # y -= 25
            total_books = 0
            total_price = 0
            row_height = 50  # Increased height for wrapped text

            for idx, ct in enumerate(cthoadon_list, start=1):
                x = left_margin
                thanh_tien = ct.SLBan * ct.GiaBan
                total_books += ct.SLBan
                total_price += thanh_tien

                # Draw row background alternating colors for better readability
                pdf.setFillColor(colors.white if idx % 2 == 0 else colors.HexColor('#F9F9F9'))
                pdf.rect(x, y - row_height, effective_width, row_height, fill=1)

                # Draw cell content
                pdf.setFillColor(colors.black)
                row_data = [
                    str(idx),
                    f"S{ct.MaSach.MaSach:03d}",
                    ct.MaSach.MaDauSach.TenSach,
                    str(ct.SLBan),
                    f"{ct.GiaBan:,}đ",
                    f"{thanh_tien:,} VNĐ"
                ]

                for data, width in zip(row_data, col_widths):
                    pdf.rect(x, y - row_height, width, row_height)
                    if width == effective_width - 350:  # Book title column
                        draw_wrapped_text(data, x, y - row_height, width, row_height)
                    else:
                        # Center align other columns
                        text_width = pdf.stringWidth(data, 'Arial Unicode', 12)
                        # pdf.drawString(x + (width - text_width)/2, y - row_height/2 + 6, data)
                        text_x = x + (width - text_width)/2
                        text_y = y - row_height + (row_height - 12)/2
                        pdf.drawString(text_x, text_y, data)
                    x += width
                y -= row_height

            # Summary section
            y -= 30
            pdf.setFont('Arial Unicode Bold', 12)
            summary_data = [
                f"Tổng Số Sách: {total_books} quyển",
                f"Tổng Tiền Sách: {total_price:,} VNĐ",
                f"Tiền Khách Trả: {hoadon.SoTienTra:,} VNĐ",
                f"Còn Lại: {hoadon.ConLai:,} VNĐ"
            ]

            for text in summary_data:
                pdf.drawString(left_margin, y, text)
                y -= 25

            pdf.save()
            return response

        except Exception as e:
            return Response({"detail": str(e)}, status=400)

class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    # Optionally add permissions